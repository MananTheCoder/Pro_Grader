from ctypes import alignment
import os
import subprocess
import time
import openpyxl
from openpyxl.styles import PatternFill

# Define folder paths (assuming current directory structure)
# base_path = './Test1'
base_path = "./"

questions_path = os.path.join(base_path, "Questions")
submissions_path = os.path.join(base_path, "Submissions")
results_path = os.path.join(base_path, "Results")

# Number of marks per question
marks_per_question = 4

# Excel color fills
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def get_test_cases(question_folder):
    """Return a list of tuples with input and expected output file paths."""
    test_cases = []
    for filename in sorted(os.listdir(question_folder)):
        if filename.endswith(".in"):
            in_file = os.path.join(question_folder, filename)
            out_file = in_file.replace(".in", ".out")
            if os.path.exists(out_file):
                test_cases.append((in_file, out_file))
    return test_cases


def run_code(user_code, input_file):
    """Run a user's code with input and return the output."""
    result = subprocess.run(
        ["g++", user_code, "-o", "temp_executable"], capture_output=True
    )
    # print(result)
    if result.returncode != 0:
        return "Compilation Error"  # Compilation error

    # Run the compiled code
    with open(input_file, "r") as infile:
        try:
            result = subprocess.run(
                "./temp_executable",
                stdin=infile,
                capture_output=True,
                text=True,
                timeout=1,
            )
            return result.stdout if result.returncode == 0 else "Runtime Error"
        except subprocess.TimeoutExpired:
            return None


def get_expected_output(expected_output_file):
    """Compare actual output with expected output file and return both expected and actual outputs."""
    with open(expected_output_file, "r") as f:
        expected_output = f.read().strip()
    return expected_output


def evaluate_user(user_folder):
    """Evaluate the user's code for all questions."""
    user_results = {}
    print("Checking code for:", user_folder)
    user_path = os.path.join(submissions_path, user_folder)

    for question in sorted(os.listdir(questions_path)):
        question_path = os.path.join(questions_path, question)
        test_cases = get_test_cases(question_path)
        passed_cases = 0
        total_cases = len(test_cases)
        failed_cases = []

        # Derive the code file based on user folder and question
        nameAsList = user_folder.lower().split("_")
        try:
            nameOfFile = (
                nameAsList[0][0] + "_" + nameAsList[1][0] + "_" + question[-1] + ".cpp"
            )
        except IndexError as e:
            print(f"Error: {e}")
            print(f"nameAsList: {nameAsList}, Skipping this user.")
            return None
        # if nameOfFile[0] == "m":
        #     continue
        code_file = os.path.join(user_path, nameOfFile)

        if not os.path.exists(code_file):
            # print("File", nameOfFile, "not found.")
            user_results[question] = {
                "passed": passed_cases,
                "total": total_cases,
                "failed_cases": "File Not Present",
            }
            continue

        for i, (input_file, expected_output_file) in enumerate(test_cases):
            current_output = run_code(code_file, input_file).strip()
            expected_output = get_expected_output(expected_output_file)
            # print(current_output, expected_output)

            if (
                not current_output.endswith("Error")
                and current_output == expected_output
            ):
                passed_cases += 1
            else:
                # Record failed test case details
                with open(input_file, "r") as f:
                    input_data = f.read().strip()
                failed_cases.append(
                    {
                        "test_case": i + 1,
                        "input": input_data,
                        "expected_output": expected_output,
                        "current_output": current_output,
                    }
                )

        user_results[question] = {
            "passed": passed_cases,
            "total": total_cases,
            "failed_cases": failed_cases,
        }

    # Store results in a text file with failure details if any
    with open(os.path.join(results_path, f"{user_folder}.txt"), "w") as f:
        for question, result in user_results.items():
            f.write(
                f"{question}: {result['passed']}/{result['total']} test cases passed.\n"
            )
            # print(result["failed_cases"])
            if result["failed_cases"]:
                if type(result["failed_cases"]) == str:
                    f.write(f"File for this code not present.\n")
                else:
                    f.write(f"Failed cases:\n")
                    for case in result["failed_cases"]:
                        f.write(f"  Test Case {case['test_case']}:\n")
                        f.write(f"    Input:           {case['input']}\n")
                        f.write(f"    Expected Output: {case['expected_output']}\n")
                        f.write(f"    Current Output:  {case['current_output']}\n\n")
    return user_results


def generate_excel_report(users_results):
    """Generate an Excel report for all users."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Test Results"

    # Headers
    ques_list = os.listdir(questions_path)
    headers = ["Name"] + ques_list + ["Total Marks"]
    sheet.append(headers)

    for row, (user, results) in enumerate(users_results.items(), start=2):
        total_marks = 0
        row_data = [user]
        for question in ques_list:
            result = results.get(question, {"passed": 0, "total": 0})
            passed, total = result["passed"], result["total"]

            row_data.append(str(passed * 100 / total) + "%")
            total_marks += (passed / total) * marks_per_question
            # if passed == total:
            #     # row_data.append("Pass")
            #     total_marks += marks_per_question
            # else:
            #     row_data.append("Fail")
        row_data.append(total_marks)
        sheet.append(row_data)

    workbook.save(os.path.join(results_path, "test_results.xlsx"))


def main():
    users_results = {}

    # Ensure results directory exists
    if not os.path.exists(results_path):
        os.makedirs(results_path)

    # temp_list = ["Sahil_Modhavadiya_53", "Sujal_Kareliya_06"]
    # for user_folder in temp_list:
    #     user_results = evaluate_user(user_folder)
    #     if user_results != None:
    #         users_results[user_folder] = user_results

    # Evaluate each user
    for user_folder in sorted(os.listdir(submissions_path)):
        user_results = evaluate_user(user_folder)
        if user_results != None:
            users_results[user_folder] = user_results

    # Generate Excel report
    generate_excel_report(users_results)
    if os.path.exists("temp_executable"):
        subprocess.run(["rm", "temp_executable"])


def temp():
    # ques_list = os.listdir(questions_path)
    # headers = ["Name"] + ques_list + ["Total Marks"]
    # user_folder = "Darshan_Padia_65 "
    print(type("1") == str)
    # user_path = os.path.join(submissions_path, user_folder)
    # nameAsList = user_folder.lower().split("_")

    # question = sorted(os.listdir(questions_path))[1]
    # question_path = os.path.join(questions_path, question)
    # test_cases = get_test_cases(question_path)

    # # for i, (input_file, expected_output_file) in enumerate(test_cases):
    # input_file = test_cases[0][0]
    # print(input_file)
    # nameAsList = user_folder.lower().split("_")
    # nameOfFile = nameAsList[0][0] + "_" + nameAsList[1][0] + "_" + question[-1] + ".cpp"
    # code_file = os.path.join(user_path, nameOfFile)
    # output = run_code(code_file, input_file)
    # print(output)
    # m1 = {1: "one", 2: "two", 3: "three"}
    # for i, (word, string_word) in enumerate(m1.items(), start=2):
    #     print(i, word, string_word)

    # quest_coll = {}
    # quest_coll['Ques1'] = {}
    # quest_coll['Ques1'][1] = {}
    # quest_coll['Ques1'][1]["in"] = 2
    # quest_coll['Ques1'][1]["out"] = "Yes"
    # run_code(cpp file, quest_coll['Ques1'][1]["in"])


if __name__ == "__main__":
    start = time.time()
    main()
    # temp()
    print(time.time() - start)
